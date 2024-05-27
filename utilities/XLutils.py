import openpyxl
from openpyxl.styles import PatternFill

def getRowCount(filepath, sheetname):
    workbook= openpyxl.load_workbook(filepath)
    sheet=workbook[sheetname]
    return sheet.max_row

def getColCount(filepath, sheetname):
    workbook= openpyxl.load_workbook(filepath)
    sheet=workbook[sheetname]
    return sheet.max_col

def readData(filepath, sheetname, rownum, colnum):
    workbook= openpyxl.load_workbook(filepath)
    sheet=workbook[sheetname]
    return sheet.cell(rownum, colnum).value

def writeData(filepath, sheetname, rownum, colnum, data):
    workbook= openpyxl.load_workbook(filepath)
    sheet=workbook[sheetname]
    sheet.cell(rownum, colnum).value= data
    workbook.save(filepath)

def fillGreenColor(filepath, sheetname, rownum, colnum):
    workbook= openpyxl.load_workbook(filepath)
    sheet=workbook[sheetname]
    greenFill= PatternFill(start_color='60b212', end_color='60b212', fill_type='solid')
    sheet.cell(rownum, colnum).fill=greenFill
    workbook.save(filepath)

def fillRedColor(filepath, sheetname, rownum, colnum):
    workbook= openpyxl.load_workbook(filepath)
    sheet=workbook[sheetname]
    redFill= PatternFill(start_color='ff0000', end_color='ff0000', fill_type='solid')
    sheet.cell(rownum, colnum).fill=redFill
    workbook.save(filepath)
